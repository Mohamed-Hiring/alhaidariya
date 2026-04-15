#!/usr/bin/env python3
"""
Regenerate import-data.js from CLEARANC 1.xlsx Sheet1.

Fixes the previous broken import that produced duplicate records
per vehicle/year. Produces exactly one record per (vehicle, year)
and prefers Completed > Paid > Pending > Failed on conflicts.
"""
import json
import re
from openpyxl import load_workbook

XLSX = "CLEARANC 1.xlsx"
OUT = "import-data.js"
NOW = "2026-04-09T00:00:00.000Z"

MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

# Sheet1 column indexes (0-based) for the year columns we care about
YEAR_COLS = [
    (12, 2025),  # 2024 to 2025 -> fiscal year 2025
    (13, 2026),  # 2025 to 2026 -> fiscal year 2026
    (14, 2027),  # 2026 to 2027 -> fiscal year 2027
]

LARGE_KEYWORDS = [
    "EXCAVATOR", "LOADER", "CRANE", "ROLLER", "TRUCK", "TANKER",
    "BUS", "FORKLIFT", "COMPRESSOR", "BULLDOZER", "DOZER",
    "GRADER", "TRAILER", "MIXER", "ASPHALT", "PAVER", "BACKHOE",
]

SUPERVISOR_NORMALIZE = {
    "ALAQIB-ASPHELT": "ALAQIB",
    "ALAQEB": "ALAQIB",
    "ALAQIB ASPHELT": "ALAQIB",
    "SAYED NAJEEB (AROON)": "SAYED NAJEEB",
    "AROON (ASPHALT)": "AROON",
    "JASIM ALGASRA": "JASIM ALGASRA",
    "JASIM  ALGASRA": "JASIM ALGASRA",
    "MUNTASER ": "MUNTASER",
    "JAHEER": "JAHIR",
}

# Status priority when two rows (shouldn't happen for Sheet1, but for safety)
# give different statuses for the same vehicle/year. Higher = wins.
STATUS_PRIORITY = {
    "completed": 4,
    "paid": 3,
    "pending_inspection": 2,
    "failed": 1,
}


def norm_supervisor(raw):
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    # Skip obvious header leftovers
    if s.upper() in {"MAME OF SUPERVISOR", "MAME", "NAME OF SUPERVISOR", "SUPERVISOR"}:
        return ""
    if s == "??":
        return ""
    key = s.upper()
    return SUPERVISOR_NORMALIZE.get(s, SUPERVISOR_NORMALIZE.get(key, s.upper()))


def norm_month(raw):
    if raw is None:
        return 0
    s = str(raw).strip().lower()
    # Sometimes cells contain more than the month word
    for k, v in MONTH_MAP.items():
        if k in s:
            return v
    return 0


def classify_type(name):
    n = (name or "").upper()
    for kw in LARGE_KEYWORDS:
        if kw in n:
            return "large"
    return "small"


def clean_str(raw):
    if raw is None:
        return ""
    return str(raw).strip()


def map_status(raw):
    """Map an Excel status cell to a (status, reason) tuple.
    Returns (None, None) if the row should not create a record.
    """
    if raw is None:
        return (None, None)
    s = str(raw).strip()
    if not s:
        return (None, None)
    u = s.upper()

    # Skip sold / cancelled
    if "SOLD" in u or "CANCEL" in u:
        return (None, None)

    # Completed (passed inspection for the season)
    if u in {"PASS", "OASS", "OK"} or u == "PASSED":
        return ("completed", "")
    if re.fullmatch(r"PASS\s*", u):
        return ("completed", "")

    # Paid (OK PAY variants)
    if "OK PAY" in u or "OK-PAY" in u or "POK PAY" in u or u == "OK-PAYMENT":
        return ("paid", "")

    # Pending inspection (appointment / needs passing)
    if "APPOINT" in u or "APPPINT" in u or "APPONTMENT" in u:
        return ("pending_inspection", "")
    if "NEED PASS" in u or u == "NEED PASSING":
        return ("pending_inspection", "")

    # Failed states (with reason)
    if "UNDER REPAIR" in u or "REPAIR" in u:
        return ("failed", s)
    if "NEEDS TIRE" in u or "WAITING TIRE" in u or "WAITAING TIRE" in u or "WAITAING PART" in u or "WAITING PART" in u:
        return ("failed", s)
    if "B-D" in u or "BD" in u or "BREAKDOWN" in u or "B/DOWN" in u:
        return ("failed", s)
    if "ACCIDENT" in u:
        return ("failed", s)
    if u == "FAIL" or u == "FAILED":
        return ("failed", s)

    # Anything else: treat as failed with the raw text as reason so it
    # shows up for the supervisor to investigate.
    return ("failed", s)


def build():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]

    vehicles = []
    records_map = {}  # (vehicle_id, year) -> record dict
    seen_keys = {}    # (plate, chassis) -> existing vehicle_id (dedupe)
    next_idx = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 16:
            continue
        sno = row[0]
        # Skip blank / summary rows
        if sno is None and not any(row):
            continue
        name = clean_str(row[2])
        if not name:
            # summary row
            continue
        # Skip stray header rows that got repeated mid-sheet
        if name.upper() == "MODEL":
            continue
        plate = clean_str(row[4])
        chassis = clean_str(row[3])
        month_raw = row[1]
        supervisor = norm_supervisor(row[15])
        notes = clean_str(row[17]) if len(row) > 17 else ""

        # Dedupe key
        key = (plate.upper(), chassis.upper())
        if key in seen_keys and plate:
            vid = seen_keys[key]
            # already have this vehicle; just merge records below
        else:
            vid = f"imp_{next_idx:04d}"
            next_idx += 1
            if plate:
                seen_keys[key] = vid
            vehicles.append({
                "id": vid,
                "name": name,
                "plateNumber": plate,
                "doorNumber": "",
                "type": classify_type(name),
                "registrationMonth": norm_month(month_raw),
                "location": "Alhaidariya",
                "chassisNumber": chassis,
                "yearOfManufacture": 0,
                "color": "",
                "driverName": "",
                "insurancePolicyNumber": "",
                "insuranceExpiryDate": "",
                "currentStatus": "",
                "assignedTo": supervisor,
                "notes": notes,
                "createdAt": NOW,
                "updatedAt": NOW,
            })

        # Walk the year columns and create at most one record per (vid, year)
        for col_idx, fiscal_year in YEAR_COLS:
            cell = row[col_idx] if col_idx < len(row) else None
            status, reason = map_status(cell)
            if status is None:
                continue
            rec_key = (vid, fiscal_year)
            existing = records_map.get(rec_key)
            if existing is not None:
                # Prefer higher-priority status when conflict
                if STATUS_PRIORITY.get(status, 0) <= STATUS_PRIORITY.get(existing["status"], 0):
                    continue
            rec = make_record(vid, fiscal_year, status, reason)
            records_map[rec_key] = rec

    records = list(records_map.values())

    # Write out
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("var IMPORT_VEHICLES = ")
        json.dump(vehicles, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
        f.write("var IMPORT_RECORDS = ")
        json.dump(records, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    # Sanity stats
    from collections import Counter
    sup_counts = Counter(v["assignedTo"] or "(unassigned)" for v in vehicles)
    status_counts = Counter(r["status"] for r in records)
    year_counts = Counter(r["year"] for r in records)

    print(f"Vehicles: {len(vehicles)}")
    print(f"Records:  {len(records)}")
    print()
    print("Supervisors:")
    for k, v in sorted(sup_counts.items(), key=lambda x: -x[1]):
        print(f"  {k:<20} {v}")
    print()
    print("Status breakdown:", dict(status_counts))
    print("Year breakdown:  ", dict(year_counts))

    # Duplicate check
    plates = [v["plateNumber"] for v in vehicles if v["plateNumber"]]
    pc = Counter(plates)
    dupes = {k: v for k, v in pc.items() if v > 1}
    if dupes:
        print(f"\nWARNING: {len(dupes)} plate duplicates remain:")
        for k, v in list(dupes.items())[:10]:
            print(f"  {k}: {v}")
    else:
        print("\nNo duplicate plates.")

    # Per-vehicle record check
    vr = Counter(r["vehicleId"] for r in records)
    max_per_v = max(vr.values()) if vr else 0
    print(f"Max records per vehicle: {max_per_v} (should be <= 3)")

    # Record duplicate check
    rec_ids = Counter(r["id"] for r in records)
    dup_ids = {k: v for k, v in rec_ids.items() if v > 1}
    if dup_ids:
        print(f"WARNING: {len(dup_ids)} duplicate record IDs")
    else:
        print("No duplicate record IDs.")


def make_record(vid, year, status, reason):
    rec = {
        "id": f"rec_{vid}_{year}",
        "vehicleId": vid,
        "year": year,
        "status": status,
        "inspectionDate": "",
        "inspectionResult": (
            "passed" if status in ("completed", "paid") else
            "failed" if status == "failed" else
            "pending"
        ),
        "failureReason": reason if status == "failed" else "",
        "failureHistory": (
            [{"date": NOW, "reason": reason, "resolvedDate": ""}]
            if status == "failed" and reason else []
        ),
        "inspectionFee": {"amount": 0, "paid": status in ("completed", "paid"), "paidDate": ""},
        "registrationFee": {"amount": 0, "paid": status == "completed", "paidDate": ""},
        "insuranceFee": {"amount": 0, "paid": status == "completed", "paidDate": ""},
        "totalCost": 0,
        "registrationCompletedDate": NOW if status == "completed" else "",
        "insuranceCompletedDate": NOW if status == "completed" else "",
        "completedDate": NOW if status == "completed" else "",
        "createdBy": "excel-import",
        "createdAt": NOW,
        "updatedAt": NOW,
    }
    return rec


if __name__ == "__main__":
    build()
